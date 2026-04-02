# -*- coding: utf-8 -*-
import io
import re
import datetime as dt
import pandas as pd
import streamlit as st

# =====================================================
# CONFIGURAÇÃO DA PÁGINA
# =====================================================
st.set_page_config(
    page_title="Comparativo Request Vs Plan",
    layout="wide"
)

st.title("Comparativo Request Vs Plan")
st.caption("Comparativo REQUEST − PLAN com filtros e resumos")

PT_BR_MESES = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]

MES_RE = re.compile(
    r'^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/\d{2}$',
    re.IGNORECASE
)

# =====================================================
# FUNÇÕES UTILITÁRIAS
# =====================================================
def _normalize_header(col):
    if isinstance(col, (pd.Timestamp, dt.date)):
        return f"{PT_BR_MESES[col.month-1]}/{col.year % 100:02d}"

    s = str(col).replace("\u00a0", " ").strip().lower()
    s = re.sub(r"[-_ ]+", "/", s)

    m = re.match(r"^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/(\d{2,4})$", s)
    if m:
        return f"{m.group(1)}/{m.group(2)[-2:]}"

    m = re.match(r"^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)(\d{2})$", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"

    return s


def detectar_colunas_mes(df):
    cols_mes = []
    debug_map = {}

    for c in df.columns:
        alias = _normalize_header(c)
        debug_map[str(c)] = alias
        if MES_RE.match(alias or ""):
            cols_mes.append(c)

    def ordem(c):
        mm, yy = debug_map[str(c)].split("/")
        return int(yy), PT_BR_MESES.index(mm)

    return sorted(cols_mes, key=ordem), debug_map


def garantir_numerico(df, meses):
    for m in meses:
        if m in df.columns:
            df[m] = pd.to_numeric(df[m], errors="coerce")
    return df


def colorir_valores(val):
    if isinstance(val, (int, float)):
        if val < 0:
            return "color:red;font-weight:bold;"
        if val > 0:
            return "color:green;font-weight:bold;"
    return ""


def formatar_tabela(df):
    df = df.fillna(0)
    cols_num = df.select_dtypes(include="number").columns

    return (
        df.style
        .format(lambda x: f"{x:,.0f}".replace(",", "."), subset=cols_num)
        .map(colorir_valores, subset=cols_num)
        .set_properties(subset=cols_num, **{"text-align": "center"})
        .set_properties(subset=df.columns.difference(cols_num),
                        **{"text-align": "left"})
    )

# =====================================================
# FUNÇÃO PRINCIPAL
# =====================================================
def gerar_passo1(xlsx_bytes, show_debug=False):

    xls_original = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")
    plan = pd.read_excel(xls_original, "PLAN", engine="openpyxl")
    req  = pd.read_excel(xls_original, "REQUEST", engine="openpyxl")

    meses_plan, map_plan = detectar_colunas_mes(plan)
    meses_req,  map_req  = detectar_colunas_mes(req)
    meses = list(dict.fromkeys(meses_plan + meses_req))

    if not meses:
        raise ValueError("Nenhuma coluna de mês encontrada.")

    plan = garantir_numerico(plan, meses)
    req  = garantir_numerico(req, meses)

    if show_debug:
        st.subheader("Diagnóstico PLAN")
        st.json(map_plan)
        st.subheader("Diagnóstico REQUEST")
        st.json(map_req)

    # =================================================
    # FILTROS
    # =================================================
    st.subheader("Filtros")

    def filtro_mult(df, col):
        if col not in df.columns:
            return None
        vals = sorted(df[col].dropna().unique())
        return st.multiselect(col, vals, default=vals)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        f_brand = filtro_mult(plan, "PRODUCT BRAND")
    with c2:
        f_market = filtro_mult(plan, "PRODUCT MARKET")
    with c3:
        f_site = filtro_mult(plan, "SITE")
    with c4:
        f_need = filtro_mult(plan, "PRODUCT NEED")

    def aplicar_filtros(df):
        if f_brand is not None:
            df = df[df["PRODUCT BRAND"].isin(f_brand)]
        if f_market is not None:
            df = df[df["PRODUCT MARKET"].isin(f_market)]
        if f_site is not None:
            df = df[df["SITE"].isin(f_site)]
        if f_need is not None:
            df = df[df["PRODUCT NEED"].isin(f_need)]
        return df

    plan = aplicar_filtros(plan)
    req  = aplicar_filtros(req)

    # =================================================
    # TABELA 1 — PRODUCT NEED (REQ - PLAN)
    # =================================================
    grp_need = ["SITE", "PRODUCT NEED"]

    plan_n = plan[grp_need + meses].groupby(grp_need, dropna=False)[meses].sum().reset_index()
    req_n  = req [grp_need + meses].groupby(grp_need, dropna=False)[meses].sum().reset_index()

    comp_n = pd.merge(plan_n, req_n, on=grp_need, how="outer",
                      suffixes=("_PLAN", "_REQ")).fillna(0)

    for m in meses:
        comp_n[m] = comp_n[f"{m}_REQ"] - comp_n[f"{m}_PLAN"]

    step1_need = comp_n[grp_need + meses].copy()
    step1_need["TOTAL"] = step1_need[meses].sum(axis=1)

    total_n = {c: "TOTAL GERAL" for c in grp_need}
    for m in meses:
        total_n[m] = step1_need[m].sum()
    total_n["TOTAL"] = step1_need["TOTAL"].sum()

    step1_need = pd.concat([step1_need, pd.DataFrame([total_n])], ignore_index=True)

    # =================================================
    # NOVA TABELA — PRODUCT NEED (SOMENTE REQUEST)
    # =================================================
    req_need = (
        req[grp_need + meses]
        .groupby(grp_need, dropna=False)[meses]
        .sum()
        .reset_index()
    )

    req_need["TOTAL"] = req_need[meses].sum(axis=1)

    total_req = {c: "TOTAL GERAL" for c in grp_need}
    for m in meses:
        total_req[m] = req_need[m].sum()
    total_req["TOTAL"] = req_need["TOTAL"].sum()

    req_need = pd.concat([req_need, pd.DataFrame([total_req])], ignore_index=True)

    # =================================================
    # TABELA 2 — ORDEM SOLICITADA
    # =================================================
    grp_serie = [
        "SITE",
        "PRODUCT NEED",
        "PRODUCT SERIES",
        "PRODUCT BRAND",
        "PRODUCT MARKET"
    ]

    plan_s = plan[grp_serie + meses].groupby(grp_serie, dropna=False)[meses].sum().reset_index()
    req_s  = req [grp_serie + meses].groupby(grp_serie, dropna=False)[meses].sum().reset_index()

    comp_s = pd.merge(plan_s, req_s, on=grp_serie, how="outer",
                      suffixes=("_PLAN", "_REQ")).fillna(0)

    for m in meses:
        comp_s[m] = comp_s[f"{m}_REQ"] - comp_s[f"{m}_PLAN"]

    step1_serie = comp_s[grp_serie + meses].copy()
    step1_serie["TOTAL"] = step1_serie[meses].sum(axis=1)

    total_s = {c: "TOTAL GERAL" for c in grp_serie}
    for m in meses:
        total_s[m] = step1_serie[m].sum()
    total_s["TOTAL"] = step1_serie["TOTAL"].sum()

    step1_serie = pd.concat([step1_serie, pd.DataFrame([total_s])], ignore_index=True)

    # =================================================
    # >>> ADIÇÃO — PRODUCT · FC · DELTA MENSAL (REQ - PLAN)
    # Mantém exatamente o padrão de colunas das abas originais (todas as
    # colunas ANTES do 1º mês, inclusive as entre DEMAND TYPE e o 1º mês)
    # =================================================
    plan_fc = plan[plan["DEMAND TYPE"] == "FC"]
    req_fc  = req [req ["DEMAND TYPE"] == "FC"]

    # Descobrir dinamicamente a posição do 1º mês no PLAN
    month_positions = [plan.columns.get_loc(c) for c in meses if c in plan.columns]
    first_month_pos = min(month_positions)
    meta_cols = list(plan.columns[:first_month_pos])  # TODAS as colunas de metadados

    # Agregar por TODAS as colunas de metadados
    plan_p = plan_fc[meta_cols + meses].groupby(meta_cols, dropna=False)[meses].sum().reset_index()
    req_p  = req_fc [meta_cols + meses].groupby(meta_cols, dropna=False)[meses].sum().reset_index()

    comp_p = pd.merge(
        plan_p, req_p,
        on=meta_cols, how="outer",
        suffixes=("_PLAN", "_REQ")
    ).fillna(0)

    # Construir a planilha no mesmo layout, com meses = DELTA (REQ − PLAN)
    step1_product_fc = comp_p[meta_cols].copy()
    for m in meses:
        step1_product_fc[m] = comp_p[f"{m}_REQ"] - comp_p[f"{m}_PLAN"]
    step1_product_fc["TOTAL"] = step1_product_fc[meses].sum(axis=1)

    # Linha TOTAL GERAL (mesma regra das demais)
    total_prod = {c: "TOTAL GERAL" for c in meta_cols}
    for m in meses:
        total_prod[m] = step1_product_fc[m].sum()
    total_prod["TOTAL"] = step1_product_fc["TOTAL"].sum()
    step1_product_fc = pd.concat([step1_product_fc, pd.DataFrame([total_prod])], ignore_index=True)

    # =================================================
    # EXPORTAR EXCEL (FORMATADO)
    # =================================================
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    buf_out = io.BytesIO()
    with pd.ExcelWriter(buf_out, engine="openpyxl") as writer:
        for sheet in xls_original.sheet_names:
            pd.read_excel(xls_original, sheet).to_excel(
                writer, sheet_name=sheet, index=False
            )

        abas = {
            "Step1_Comparativo_Serie": step1_serie,
            "Step1_Comparativo_Need": step1_need,
            "Resumo_Request_Product_Need": req_need,
            # Nova aba com TODAS as colunas de metadados preservadas
            "REQ x PLAN FC - Produto Mensal": step1_product_fc,
        }

        for nome, df in abas.items():
            df.to_excel(writer, sheet_name=nome, index=False)
            ws = writer.book[nome]

            cols_num = df.select_dtypes(include="number").columns
            idx_cols = [df.columns.get_loc(c) + 1 for c in cols_num]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for idx in idx_cols:
                    cell = row[idx - 1]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                        if cell.value < 0:
                            cell.font = Font(color="FF0000", bold=True)
                        elif cell.value > 0:
                            cell.font = Font(color="008000", bold=True)

                if str(row[0].value).upper() == "TOTAL GERAL":
                    for cell in row:
                        cell.font = Font(bold=True)

            for col in ws.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0,
                    *[0]
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    return buf_out.getvalue(), step1_serie, step1_need, req_need

# =====================================================
# UI
# =====================================================
uploaded = st.file_uploader("Envie o Excel (PLAN e REQUEST)", type=["xlsx"])
debug = st.checkbox("Exibir diagnóstico", value=False)

if uploaded:
    excel_out, df_serie, df_need, df_req_need = gerar_passo1(uploaded.read(), debug)

    st.subheader("Comparativo por PRODUCT NEED + PRODUCT SERIES")
    st.dataframe(formatar_tabela(df_serie), use_container_width=True)

    st.subheader("Resumo por PRODUCT NEED (REQ - PLAN)")
    st.dataframe(formatar_tabela(df_need), use_container_width=True)

    st.subheader("Resumo por PRODUCT NEED (REQUEST)")
    st.dataframe(formatar_tabela(df_req_need), use_container_width=True)

    st.download_button(
        "⬇️ Baixar Excel",
        data=excel_out,
        file_name="saida_step1.xlsx"
    )
else:
    st.info("Faça upload do Excel para iniciar.")
